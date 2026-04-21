# -*- coding: utf-8 -*-

import sqlite3
import os
import io
import csv
import json
import hashlib
import base64
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, session, Response, redirect

from plant_model import (
    identify_plant,
    get_all_plants_arabic,
    get_plant_catalog,
    get_recommendations,
    get_cities,
    get_home_types,
    get_tip_of_day,
    invalidate_catalog_cache,
)
from catalog_db import (
    create_catalog_tables,
    migrate_catalog_plant_columns,
    migrate_after_plants_and_care_logs,
    seed_cities_and_catalog_if_empty,
    backfill_catalog_plant_fields,
    migrate_users_city_id,
    resolve_city_id,
)


def _parse_dt(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    if len(s) >= 10:
        try:
            return datetime.strptime(s[:10], '%Y-%m-%d')
        except ValueError:
            pass
    return None


def _plant_health_score(p, now):
    wid = int(p.get('watering_interval_days') or 7)
    fid = int(p.get('fertilizing_interval_days') or 30)
    t_lw = _parse_dt(p.get('last_watering_date'))
    t_lf = _parse_dt(p.get('last_fertilizing_date'))
    score = 100.0
    if t_lw:
        days_since_w = (now - t_lw).days
        if days_since_w > wid:
            overdue = days_since_w - wid
            score -= min(55.0, overdue * 8.0)
    else:
        score -= 12.0
    if t_lf:
        days_since_f = (now - t_lf).days
        if days_since_f > fid:
            overdue_f = days_since_f - fid
            score -= min(35.0, overdue_f * 5.0)
    else:
        score -= 12.0
    return int(max(0, min(100, round(score))))


def _health_status_ar(score):
    if score >= 82:
        return 'ممتاز'
    if score >= 65:
        return 'جيد'
    if score >= 45:
        return 'يحتاج متابعة'
    return 'ضعيف'


def _care_streak_days(conn, user_id):
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT DISTINCT substr(action_date, 1, 10) AS d
        FROM care_logs cl JOIN plants p ON cl.plant_id = p.id
        WHERE p.user_id = ?
        ''',
        (user_id,),
    )
    rows = [r[0] for r in cursor.fetchall()]
    date_set = set()
    for r in rows:
        try:
            date_set.add(datetime.strptime(r, '%Y-%m-%d').date())
        except ValueError:
            pass
    if not date_set:
        return 0
    from datetime import date as date_cls
    today = date_cls.today()
    streak = 0
    d = today
    while d in date_set:
        streak += 1
        d -= timedelta(days=1)
    return streak


app = Flask(__name__, static_folder='.')
app.secret_key = os.environ.get('FLORABIT_SECRET_KEY', 'florabit-dev-change-in-production')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
if os.environ.get('RENDER') or os.environ.get('FLORABIT_SECURE_COOKIES', '').lower() in (
    '1',
    'true',
    'yes',
):
    app.config['SESSION_COOKIE_SECURE'] = True


def _cors_allowed_origins():
    allowed = {
        'http://localhost:5000',
        'http://127.0.0.1:5000',
        'http://localhost:5500',
        'http://127.0.0.1:5500',
    }
    render_public = (os.environ.get('RENDER_EXTERNAL_URL') or '').strip().rstrip('/')
    if render_public:
        allowed.add(render_public)
    extra = os.environ.get('FLORABIT_CORS_ORIGINS', '')
    for part in extra.split(','):
        o = part.strip().rstrip('/')
        if o:
            allowed.add(o)
    return allowed


def _cors(response):
    origin = request.headers.get('Origin')
    allowed = _cors_allowed_origins()
    if origin in allowed:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    else:
        response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, PATCH, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-User-Id'
    return response


@app.after_request
def add_cors_headers(response):
    return _cors(response)


def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': 'يتطلب صلاحية مدير النظام'}), 403
        return f(*args, **kwargs)
    return wrapped


def _account_block_reason(row_dict):
    if int(row_dict.get('is_active') or 1) == 0:
        return 'تم إيقاف هذا الحساب'
    su = row_dict.get('suspended_until')
    if su:
        end = _parse_dt(str(su))
        if end and datetime.now() < end:
            return 'الحساب معلّق حتى ' + str(su)[:19]
    return None


def _plant_access_denied(plant_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM plants WHERE id = ?', (plant_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'غير موجود'}), 404
    owner = row[0]
    sid = session.get('user_id')
    role = session.get('role')
    hdr = request.headers.get('X-User-Id')
    ok = role == 'admin' or (sid and sid == owner) or (hdr is not None and str(hdr) == str(owner))
    if not ok:
        return jsonify({'error': 'غير مصرّح'}), 403
    return None


def _ensure_admin_exists(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users WHERE COALESCE(role,'user') = 'admin'")
    n = cursor.fetchone()[0]
    if n > 0:
        return
    env_email = (os.environ.get('FLORABIT_ADMIN_EMAIL') or '').strip().lower()
    if env_email:
        cursor.execute(
            "UPDATE users SET role = 'admin' WHERE lower(email) = ?",
            (env_email,),
        )
    else:
        cursor.execute(
            "UPDATE users SET role = 'admin' WHERE id = (SELECT MIN(id) FROM users)"
        )
    conn.commit()

DATABASE = os.path.join(os.path.dirname(__file__), 'database.db')


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT,
            city TEXT,
            home_type TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    for col in ['password', 'city', 'home_type', 'created_at', 'updated_at', 'role', 'suspended_until']:
        try:
            cursor.execute(f'ALTER TABLE users ADD COLUMN {col} TEXT')
        except sqlite3.OperationalError:
            pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN is_active INTEGER DEFAULT 1')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN terms_privacy_accepted_at TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute(
            'ALTER TABLE users ADD COLUMN plant_public_scan_consent INTEGER DEFAULT 0'
        )
    except sqlite3.OperationalError:
        pass

    create_catalog_tables(cursor)
    migrate_catalog_plant_columns(cursor)
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN city_id INTEGER REFERENCES cities(id)')
    except sqlite3.OperationalError:
        pass

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS plants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            catalog_id INTEGER,
            name TEXT NOT NULL,
            type TEXT NOT NULL,
            indoor_outdoor TEXT,
            watering_interval_days INTEGER NOT NULL,
            fertilizing_interval_days INTEGER NOT NULL,
            last_watering_date TEXT,
            last_fertilizing_date TEXT,
            notes TEXT,
            created_at TEXT,
            updated_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN catalog_id INTEGER')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN indoor_outdoor TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN notes TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN created_at TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN updated_at TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN latitude REAL')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE plants ADD COLUMN longitude REAL')
    except sqlite3.OperationalError:
        pass
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS care_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plant_id INTEGER NOT NULL,
            action_type TEXT NOT NULL,
            action_date TEXT NOT NULL,
            notes TEXT,
            FOREIGN KEY (plant_id) REFERENCES plants(id)
        )
    ''')
    migrate_after_plants_and_care_logs(cursor)

    conn.commit()
    conn.close()
    conn = get_db()
    try:
        seed_cities_and_catalog_if_empty(conn)
        backfill_catalog_plant_fields(conn)
        migrate_users_city_id(conn)
        invalidate_catalog_cache()
    except Exception:
        pass
    conn.close()
    conn = get_db()
    _ensure_admin_exists(conn)
    conn.close()


# ============ API Endpoints ============

@app.route('/api/<path:subpath>', methods=['OPTIONS'])
def options_handler(subpath):
    return '', 204


@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT id, name, email, city, home_type, role, is_active, suspended_until, created_at,
                  terms_privacy_accepted_at, plant_public_scan_consent
           FROM users ORDER BY id'''
    )
    users = [dict(row) for row in cursor.fetchall()]
    conn.close()
    for u in users:
        if u.get('role') is None:
            u['role'] = 'user'
        if u.get('is_active') is None:
            u['is_active'] = 1
        if u.get('plant_public_scan_consent') is None:
            u['plant_public_scan_consent'] = 0
    return jsonify(users)


@app.route('/api/users', methods=['POST'])
def create_user():
    data = request.get_json()
    if not data or not data.get('name') or not data.get('email'):
        return jsonify({'error': 'الاسم والبريد مطلوبان'}), 400
    password = data.get('password', '')
    if not password or len(password) < 4:
        return jsonify({'error': 'كلمة المرور يجب أن تكون 4 أحرف على الأقل'}), 400
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    city = data.get('city', '') or ''
    home_type = data.get('home_type', '') or ''
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if session.get('role') == 'admin':
        role = data.get('role', 'user')
        if role not in ('user', 'admin'):
            role = 'user'
    else:
        role = 'user'
    conn = get_db()
    cursor = conn.cursor()
    city_id = resolve_city_id(cursor, city) if city else None
    try:
        cursor.execute(
            '''INSERT INTO users (name, email, password, city, home_type, created_at, updated_at, role, is_active, city_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)''',
            (data['name'], data['email'], pw_hash, city, home_type, now, now, role, city_id),
        )
        conn.commit()
        user_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': user_id, 'name': data['name'], 'email': data['email'], 'role': role}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'البريد مستخدم مسبقاً'}), 400


@app.route('/api/users/<int:user_id>', methods=['GET'])
def get_user(user_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT id, name, email, city, home_type, terms_privacy_accepted_at,
                  plant_public_scan_consent FROM users WHERE id = ?''',
        (user_id,),
    )
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'User not found'}), 404
    return jsonify(dict(row))


@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    data = request.get_json()
    sid = session.get('user_id')
    role = session.get('role')
    hdr = request.headers.get('X-User-Id')
    if role != 'admin' and sid != user_id and (hdr is None or str(hdr) != str(user_id)):
        return jsonify({'error': 'غير مصرّح بتعديل هذا المستخدم'}), 403
    conn = get_db()
    cursor = conn.cursor()
    updates = []
    values = []
    for f in ['name', 'email', 'city', 'home_type', 'terms_privacy_accepted_at']:
        if f in data:
            updates.append(f'{f} = ?')
            values.append(data[f])
    if 'plant_public_scan_consent' in data:
        v = data['plant_public_scan_consent']
        updates.append('plant_public_scan_consent = ?')
        values.append(1 if v in (True, 1, '1', 'true') else 0)
    if 'city' in data:
        cid = resolve_city_id(cursor, data.get('city'))
        updates.append('city_id = ?')
        values.append(cid)
    if session.get('role') == 'admin' and 'role' in data:
        r = data['role']
        if r in ('user', 'admin'):
            cursor.execute("SELECT COUNT(*) FROM users WHERE COALESCE(role,'user') = 'admin'")
            ac = cursor.fetchone()[0]
            cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,))
            old = cursor.fetchone()
            old_r = (old[0] or 'user') if old else 'user'
            if old_r == 'admin' and r == 'user' and ac <= 1:
                conn.close()
                return jsonify({'error': 'لا يمكن إزالة صلاحية آخر مدير'}), 400
            updates.append('role = ?')
            values.append(r)
    if not updates:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    updates.append('updated_at = ?')
    values.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    values.append(user_id)
    cursor.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data or not data.get('email'):
        return jsonify({'error': 'البريد مطلوب'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT id, name, email, password, city, home_type, role, is_active, suspended_until,
                  terms_privacy_accepted_at, plant_public_scan_consent
           FROM users WHERE email = ?''',
        (data['email'],),
    )
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'المستخدم غير موجود'}), 404
    row_dict = dict(row)
    pw = row_dict.pop('password', None)
    for k in [
        'city',
        'home_type',
        'role',
        'is_active',
        'suspended_until',
        'terms_privacy_accepted_at',
        'plant_public_scan_consent',
    ]:
        if k not in row_dict:
            row_dict[k] = None
    if row_dict.get('plant_public_scan_consent') is None:
        row_dict['plant_public_scan_consent'] = 0
    if row_dict.get('role') is None:
        row_dict['role'] = 'user'
    if row_dict.get('is_active') is None:
        row_dict['is_active'] = 1
    # تحقق من كلمة المرور أولاً حتى لا يُستنتج حالة الحساب دون إثبات الهوية
    if pw is not None and pw != '':
        given = data.get('password', '')
        pw_hash = hashlib.sha256(given.encode()).hexdigest()
        if pw_hash != pw:
            return jsonify({'error': 'كلمة المرور غير صحيحة'}), 401
    block = _account_block_reason(row_dict)
    if block:
        return jsonify({'error': block}), 403
    session['user_id'] = row_dict['id']
    session['role'] = row_dict.get('role') or 'user'
    session.permanent = True
    out = {k: v for k, v in row_dict.items() if k != 'suspended_until'}
    return jsonify(out)


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'logged_in': False}), 200
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT id, name, email, city, home_type, role, is_active, suspended_until
           FROM users WHERE id = ?''',
        (uid,),
    )
    row = cursor.fetchone()
    conn.close()
    if row is None:
        session.clear()
        return jsonify({'logged_in': False}), 200
    d = dict(row)
    if d.get('role') is None:
        d['role'] = 'user'
    br = _account_block_reason(d)
    if br:
        session.clear()
        return jsonify({'logged_in': False, 'error': br}), 200
    d.pop('password', None)
    return jsonify({'logged_in': True, 'user': d, 'role': d.get('role') or 'user'})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    sid = session.get('user_id')
    if sid == user_id:
        return jsonify({'error': 'لا يمكنك حذف حسابك وأنت مسجّل الدخول'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users WHERE COALESCE(role,'user') = 'admin'")
    admin_count = cursor.fetchone()[0]
    cursor.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    r = cursor.fetchone()
    if r and (r[0] or 'user') == 'admin' and admin_count <= 1:
        conn.close()
        return jsonify({'error': 'لا يمكن حذف آخر مدير في النظام'}), 400
    cursor.execute('SELECT id FROM plants WHERE user_id = ?', (user_id,))
    pids = [x[0] for x in cursor.fetchall()]
    for pid in pids:
        cursor.execute('DELETE FROM care_logs WHERE plant_id = ?', (pid,))
        cursor.execute('DELETE FROM plants WHERE id = ?', (pid,))
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>/suspend', methods=['POST'])
@admin_required
def suspend_user(user_id):
    if session.get('user_id') == user_id:
        return jsonify({'error': 'لا يمكن تعليق حسابك'}), 400
    data = request.get_json() or {}
    days = int(data.get('days', 30))
    end = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'UPDATE users SET is_active = 0, suspended_until = ?, updated_at = ? WHERE id = ?',
        (end, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_id),
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'suspended_until': end})


@app.route('/api/users/<int:user_id>/activate', methods=['POST'])
@admin_required
def activate_user(user_id):
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(
        'UPDATE users SET is_active = 1, suspended_until = NULL, updated_at = ? WHERE id = ?',
        (now, user_id),
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/plants', methods=['GET'])
def get_plants():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        if session.get('role') != 'admin':
            return jsonify({'error': 'مرفوض: حدّد user_id أو سجّل دخول كمدير'}), 403
    conn = get_db()
    cursor = conn.cursor()
    if user_id:
        cursor.execute(
            '''SELECT p.*, u.name as user_name FROM plants p
               JOIN users u ON p.user_id = u.id WHERE p.user_id = ?''',
            (user_id,),
        )
    else:
        cursor.execute(
            '''SELECT p.*, u.name as user_name FROM plants p
               JOIN users u ON p.user_id = u.id'''
        )
    plants = [dict(row) for row in cursor.fetchall()]
    conn.close()
    for p in plants:
        if 'user_name' in p:
            del p['user_name']
    return jsonify(plants)


@app.route('/api/plants', methods=['POST'])
def create_plant():
    data = request.get_json()
    required = ['user_id', 'name', 'type', 'watering_interval_days', 'fertilizing_interval_days']
    if not all(data.get(k) for k in required):
        return jsonify({'error': 'Missing required fields'}), 400
    sid = session.get('user_id')
    if sid and session.get('role') != 'admin':
        if int(data.get('user_id')) != int(sid):
            return jsonify({'error': 'لا يمكن إضافة نبتة لمستخدم آخر'}), 403
    indoor_outdoor = data.get('indoor_outdoor') or ('داخلي' if data.get('type') == 'داخلي' else 'خارجي')
    catalog_id = data.get('catalog_id')
    notes = data.get('notes', '')

    def _parse_coord(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    lat = _parse_coord(data.get('latitude'))
    lng = _parse_coord(data.get('longitude'))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO plants (user_id, catalog_id, name, type, indoor_outdoor, 
                     watering_interval_days, fertilizing_interval_days, notes, latitude, longitude, created_at, updated_at) 
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (data['user_id'], catalog_id, data['name'], data['type'], indoor_outdoor,
                    data['watering_interval_days'], data['fertilizing_interval_days'], notes, lat, lng, now, now))
    conn.commit()
    plant_id = cursor.lastrowid
    conn.close()
    return jsonify({'id': plant_id, 'indoor_outdoor': indoor_outdoor, **{k: data[k] for k in required}}), 201


@app.route('/api/plants/<int:plant_id>', methods=['GET'])
def get_plant(plant_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM plants WHERE id = ?', (plant_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'Plant not found'}), 404
    plant = dict(row)
    owner_id = plant.get('user_id')
    cursor.execute(
        'SELECT plant_public_scan_consent FROM users WHERE id = ?',
        (owner_id,),
    )
    urow = cursor.fetchone()
    conn.close()
    public_ok = urow and int(urow[0] or 0) == 1
    sid = session.get('user_id')
    role = session.get('role')
    hdr = request.headers.get('X-User-Id')
    is_owner = (hdr is not None and str(hdr) == str(owner_id)) or (
        sid is not None and int(sid) == int(owner_id)
    )
    is_admin = role == 'admin'
    if is_owner or is_admin or public_ok:
        return jsonify(plant)
    return jsonify({
        'error': 'صاحب النبتة لم يفعّل مشاركة البيانات عند مسح الباركود.',
        'code': 'plant_scan_consent_required',
    }), 403


@app.route('/api/plants/<int:plant_id>', methods=['PUT'])
def update_plant(plant_id):
    data = request.get_json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM plants WHERE id = ?', (plant_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'غير موجود'}), 404
    owner = row[0]
    sid = session.get('user_id')
    role = session.get('role')
    hdr = request.headers.get('X-User-Id')
    ok = role == 'admin' or (sid and sid == owner) or (hdr is not None and str(hdr) == str(owner))
    if not ok:
        conn.close()
        return jsonify({'error': 'غير مصرّح'}), 403
    fields = ['name', 'type', 'indoor_outdoor', 'watering_interval_days', 'fertilizing_interval_days', 'notes']

    def _parse_coord(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    updates = []
    values = []
    for f in fields:
        if f in data:
            updates.append(f'{f} = ?')
            values.append(data[f])
    if 'latitude' in data:
        updates.append('latitude = ?')
        values.append(_parse_coord(data.get('latitude')))
    if 'longitude' in data:
        updates.append('longitude = ?')
        values.append(_parse_coord(data.get('longitude')))
    if not updates:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    values.append(plant_id)
    cursor.execute(f"UPDATE plants SET {', '.join(updates)} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/plants/<int:plant_id>', methods=['DELETE'])
def delete_plant(plant_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM plants WHERE id = ?', (plant_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'غير موجود'}), 404
    owner = row[0]
    sid = session.get('user_id')
    role = session.get('role')
    hdr = request.headers.get('X-User-Id')
    ok = False
    if role == 'admin':
        ok = True
    elif sid and sid == owner:
        ok = True
    elif hdr is not None and str(hdr) == str(owner):
        ok = True
    if not ok:
        conn.close()
        return jsonify({'error': 'غير مصرّح بحذف هذه النبتة'}), 403
    cursor.execute('DELETE FROM care_logs WHERE plant_id = ?', (plant_id,))
    cursor.execute('DELETE FROM plants WHERE id = ?', (plant_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/plants/<int:plant_id>/water', methods=['POST'])
def water_plant(plant_id):
    denied = _plant_access_denied(plant_id)
    if denied is not None:
        return denied
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(
        'INSERT INTO care_logs (plant_id, action_type, action_date, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
        (plant_id, 'watering', now, request.get_json().get('notes', '') if request.get_json() else '', now, now),
    )
    cursor.execute('UPDATE plants SET last_watering_date = ? WHERE id = ?', (now, plant_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'last_watering_date': now})


@app.route('/api/plants/<int:plant_id>/fertilize', methods=['POST'])
def fertilize_plant(plant_id):
    denied = _plant_access_denied(plant_id)
    if denied is not None:
        return denied
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(
        'INSERT INTO care_logs (plant_id, action_type, action_date, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
        (plant_id, 'fertilizing', now, request.get_json().get('notes', '') if request.get_json() else '', now, now),
    )
    cursor.execute('UPDATE plants SET last_fertilizing_date = ? WHERE id = ?', (now, plant_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'last_fertilizing_date': now})


@app.route('/api/plants/<int:plant_id>/light', methods=['POST'])
def light_log_plant(plant_id):
    denied = _plant_access_denied(plant_id)
    if denied is not None:
        return denied
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    notes = ''
    if request.get_json():
        notes = request.get_json().get('notes', '') or ''
    cursor.execute(
        'INSERT INTO care_logs (plant_id, action_type, action_date, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
        (plant_id, 'lighting', now, notes, now, now),
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'action_date': now})


@app.route('/api/identify-plant', methods=['POST'])
def api_identify_plant():
    if 'image' not in request.files and not request.get_json():
        return jsonify({'error': 'الصورة مطلوبة'}), 400
    img_data = None
    if 'image' in request.files:
        f = request.files['image']
        if f and f.filename:
            img_data = f.read()
    elif request.is_json:
        data = request.get_json()
        b64 = data.get('image_base64', '')
        if b64:
            img_data = base64.b64decode(b64.split(',')[-1] if ',' in b64 else b64)
    if not img_data:
        return jsonify({'error': 'الصورة مطلوبة'}), 400
    results = identify_plant(img_data)
    return jsonify({'predictions': results})


@app.route('/api/plants/arabic-list', methods=['GET'])
def api_plants_arabic_list():
    plants = get_all_plants_arabic()
    return jsonify(plants)


@app.route('/api/plant-catalog', methods=['GET'])
def api_plant_catalog():
    catalog = get_plant_catalog()
    return jsonify(catalog)


@app.route('/api/plants/recommendations', methods=['GET'])
def api_plant_recommendations():
    city = request.args.get('city', '')
    home_type = request.args.get('home_type', '')
    recommendations = get_recommendations(city or None, home_type or None)
    return jsonify(recommendations)


@app.route('/api/cities', methods=['GET'])
def api_cities():
    return jsonify(get_cities())


@app.route('/api/home-types', methods=['GET'])
def api_home_types():
    return jsonify(get_home_types())


@app.route('/api/admin/cities', methods=['GET', 'POST'])
@admin_required
def admin_cities_manage():
    conn = get_db()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute('SELECT id, name, sort_order FROM cities ORDER BY sort_order, id')
        rows = [{'id': r[0], 'name': r[1], 'sort_order': r[2]} for r in cursor.fetchall()]
        conn.close()
        return jsonify(rows)
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        conn.close()
        return jsonify({'error': 'name مطلوب'}), 400
    so = int(data.get('sort_order', 999))
    try:
        cursor.execute(
            'INSERT INTO cities (name, sort_order) VALUES (?, ?)',
            (name, so),
        )
        conn.commit()
        cid = cursor.lastrowid
        conn.close()
        invalidate_catalog_cache()
        return jsonify({'id': cid, 'name': name, 'sort_order': so}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'المدينة موجودة'}), 400


@app.route('/api/admin/cities/<int:city_id>', methods=['DELETE'])
@admin_required
def admin_delete_city(city_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM users WHERE city_id = ?', (city_id,))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return jsonify({'error': 'يوجد مستخدمون مرتبطون بهذه المدينة'}), 400
    cursor.execute('DELETE FROM catalog_plant_cities WHERE city_id = ?', (city_id,))
    cursor.execute('DELETE FROM cities WHERE id = ?', (city_id,))
    conn.commit()
    conn.close()
    invalidate_catalog_cache()
    return jsonify({'success': True})


@app.route('/api/admin/catalog-plant-cities', methods=['POST', 'DELETE'])
@admin_required
def admin_catalog_plant_cities():
    conn = get_db()
    cursor = conn.cursor()
    if request.method == 'POST':
        data = request.get_json() or {}
        pid = data.get('catalog_plant_id')
        cid = data.get('city_id')
        if pid is None or cid is None:
            conn.close()
            return jsonify({'error': 'catalog_plant_id و city_id مطلوبان'}), 400
        cursor.execute(
            'INSERT OR IGNORE INTO catalog_plant_cities (catalog_plant_id, city_id) VALUES (?, ?)',
            (int(pid), int(cid)),
        )
        conn.commit()
        conn.close()
        invalidate_catalog_cache()
        return jsonify({'success': True}), 201
    pid = request.args.get('catalog_plant_id', type=int)
    cid = request.args.get('city_id', type=int)
    if pid is None or cid is None:
        conn.close()
        return jsonify({'error': 'مرّر catalog_plant_id و city_id في الاستعلام'}), 400
    cursor.execute(
        'DELETE FROM catalog_plant_cities WHERE catalog_plant_id = ? AND city_id = ?',
        (pid, cid),
    )
    conn.commit()
    conn.close()
    invalidate_catalog_cache()
    return jsonify({'success': True})


@app.route('/api/admin/catalog-plants', methods=['GET'])
@admin_required
def admin_catalog_plants_list():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT id, name, type, indoor_outdoor, watering, fertilizing, light FROM catalog_plants ORDER BY id'
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route('/api/admin/stats', methods=['GET'])
@admin_required
def admin_stats():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM users')
    users_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM plants')
    plants_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM care_logs')
    logs_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM users WHERE COALESCE(is_active,1) = 0')
    suspended_count = cursor.fetchone()[0]
    conn.close()
    return jsonify({
        'users': users_count,
        'plants': plants_count,
        'care_logs': logs_count,
        'suspended_or_inactive': suspended_count,
    })


@app.route('/api/admin/chart-insights', methods=['GET'])
@admin_required
def admin_chart_insights():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT action_type, COUNT(*) FROM care_logs GROUP BY action_type'
    )
    by_action = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute(
        '''SELECT date(action_date) AS d, COUNT(*) AS c FROM care_logs
           WHERE date(action_date) >= date('now', '-13 days')
           GROUP BY date(action_date) ORDER BY d'''
    )
    daily = [{'date': row[0], 'count': row[1]} for row in cursor.fetchall()]
    conn.close()
    return jsonify({'by_action': by_action, 'daily': daily})


def _fetch_admin_report_bundle():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM users')
    uc = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM plants')
    pc = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM care_logs')
    lc = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM users WHERE COALESCE(is_active,1) = 0')
    sc = cursor.fetchone()[0]
    cursor.execute(
        '''SELECT id, name, email, city, home_type, COALESCE(role,'user') AS role,
           COALESCE(is_active,1) AS is_active, created_at FROM users ORDER BY id'''
    )
    users_rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute(
        '''SELECT id, user_id, catalog_id, name, type, indoor_outdoor,
           watering_interval_days, fertilizing_interval_days, last_watering_date, last_fertilizing_date,
           notes, created_at FROM plants ORDER BY id'''
    )
    plant_rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute(
        'SELECT id, plant_id, action_type, action_date, notes FROM care_logs ORDER BY action_date DESC LIMIT 5000'
    )
    log_rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute('SELECT action_type, COUNT(*) FROM care_logs GROUP BY action_type')
    by_action = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute(
        '''SELECT date(action_date) AS d, COUNT(*) AS c FROM care_logs
           WHERE date(action_date) >= date('now', '-13 days')
           GROUP BY date(action_date) ORDER BY d'''
    )
    daily = [{'date': row[0], 'count': row[1]} for row in cursor.fetchall()]
    conn.close()
    return {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'stats': {
            'users': uc,
            'plants': pc,
            'care_logs': lc,
            'suspended_or_inactive': sc,
        },
        'users': users_rows,
        'plants': plant_rows,
        'care_logs': log_rows,
        'chart_insights': {'by_action': by_action, 'daily': daily},
    }


def _sqlesc_sql(val):
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''").replace('\\', '\\\\') + "'"


def _bundle_to_txt(bundle):
    lines = [
        'تقرير فلورابيت — نسخة نصية',
        'تاريخ التوليد: ' + bundle['generated_at'],
        '',
        '— ملخص —',
    ]
    for k, v in bundle['stats'].items():
        lines.append('  %s: %s' % (k, v))
    lines.extend(['', '— المستخدمون (%d) —' % len(bundle['users'])])
    for u in bundle['users']:
        lines.append('  [%s] %s | %s | %s | %s' % (
            u.get('id'), u.get('name'), u.get('email'), u.get('city') or '—', u.get('home_type') or '—',
        ))
    lines.extend(['', '— عيّنة نباتات (أول 40) —'])
    for p in bundle['plants'][:40]:
        lines.append('  plant_id=%s user_id=%s name=%s' % (p.get('id'), p.get('user_id'), p.get('name')))
    lines.extend(['', '— عيّنة سجلات العناية (أول 50) —'])
    for g in bundle['care_logs'][:50]:
        lines.append('  %s | plant=%s | %s | %s' % (
            g.get('action_date'), g.get('plant_id'), g.get('action_type'), g.get('notes') or '',
        ))
    return '\n'.join(lines)


def _bundle_to_sql(bundle, mysql_style=False):
    hdr = '-- Florabit SQL dump (%s)\n-- %s\nSET NAMES utf8mb4;\n\n' % (
        'MySQL/MariaDB' if mysql_style else 'SQLite',
        bundle['generated_at'],
    )
    parts = [hdr]
    tbl_users = '`users`' if mysql_style else 'users'
    tbl_plants = '`plants`' if mysql_style else 'plants'
    tbl_logs = '`care_logs`' if mysql_style else 'care_logs'
    for u in bundle['users']:
        cols = ('id', 'name', 'email', 'city', 'home_type', 'role', 'is_active', 'created_at')
        vals = ', '.join(_sqlesc_sql(u.get(c)) for c in cols)
        parts.append('INSERT INTO %s (%s) VALUES (%s);\n' % (tbl_users, ', '.join(cols), vals))
    for p in bundle['plants']:
        cols = (
            'id', 'user_id', 'catalog_id', 'name', 'type', 'indoor_outdoor',
            'watering_interval_days', 'fertilizing_interval_days', 'last_watering_date', 'last_fertilizing_date',
            'notes', 'created_at',
        )
        vals = ', '.join(_sqlesc_sql(p.get(c)) for c in cols)
        parts.append('INSERT INTO %s (%s) VALUES (%s);\n' % (tbl_plants, ', '.join(cols), vals))
    for g in bundle['care_logs']:
        cols = ('id', 'plant_id', 'action_type', 'action_date', 'notes')
        vals = ', '.join(_sqlesc_sql(g.get(c)) for c in cols)
        parts.append('INSERT INTO %s (%s) VALUES (%s);\n' % (tbl_logs, ', '.join(cols), vals))
    return ''.join(parts)


def _bundle_to_csv_bytes(bundle):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['section', 'key', 'value'])
    for k, v in bundle['stats'].items():
        w.writerow(['stats', k, v])
    w.writerow([])
    w.writerow(['users'])
    if bundle['users']:
        uk = list(bundle['users'][0].keys())
        w.writerow(uk)
        for row in bundle['users']:
            w.writerow([row.get(x) for x in uk])
    w.writerow([])
    w.writerow(['plants'])
    if bundle['plants']:
        pk = list(bundle['plants'][0].keys())
        w.writerow(pk)
        for row in bundle['plants']:
            w.writerow([row.get(x) for x in pk])
    w.writerow([])
    w.writerow(['care_logs'])
    if bundle['care_logs']:
        gk = list(bundle['care_logs'][0].keys())
        w.writerow(gk)
        for row in bundle['care_logs']:
            w.writerow([row.get(x) for x in gk])
    return '\ufeff' + buf.getvalue()


def _bundle_to_xlsx_bytes(bundle):
    try:
        from openpyxl import Workbook
    except ImportError:
        return None
    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'stats'
    ws0.append(['metric', 'value'])
    for k, v in bundle['stats'].items():
        ws0.append([k, v])
    ws1 = wb.create_sheet('users')
    if bundle['users']:
        uk = list(bundle['users'][0].keys())
        ws1.append(uk)
        for row in bundle['users']:
            ws1.append([row.get(x) for x in uk])
    ws2 = wb.create_sheet('plants')
    if bundle['plants']:
        pk = list(bundle['plants'][0].keys())
        ws2.append(pk)
        for row in bundle['plants']:
            ws2.append([row.get(x) for x in pk])
    ws3 = wb.create_sheet('care_logs')
    if bundle['care_logs']:
        gk = list(bundle['care_logs'][0].keys())
        ws3.append(gk)
        for row in bundle['care_logs']:
            ws3.append([row.get(x) for x in gk])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _pdf_ascii(val):
    if val is None:
        return ''
    return ''.join((c if ord(c) < 128 else '?') for c in str(val))


def _bundle_to_pdf_bytes(bundle):
    try:
        from fpdf import FPDF
    except ImportError:
        return None
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_font('Helvetica', '', 10)
    pdf.add_page()
    pw = pdf.w - pdf.l_margin - pdf.r_margin
    title = 'Florabit report ' + _pdf_ascii(bundle['generated_at'])
    pdf.multi_cell(pw, 6, title)
    pdf.ln(3)
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(
        pw, 5,
        'Stats: users=%s plants=%s care_logs=%s suspended=%s' % (
            bundle['stats']['users'], bundle['stats']['plants'], bundle['stats']['care_logs'],
            bundle['stats']['suspended_or_inactive'],
        ),
    )
    pdf.ln(2)
    pdf.multi_cell(pw, 5, 'Users (first 25, Arabic as ? in PDF):')
    for u in bundle['users'][:25]:
        line = '#%s %s <%s> %s' % (
            u.get('id'),
            _pdf_ascii(u.get('name')),
            _pdf_ascii(u.get('email')),
            _pdf_ascii(u.get('city') or '-'),
        )
        pdf.multi_cell(pw, 5, line)
    out = pdf.output(dest='S')
    if isinstance(out, str):
        return out.encode('latin1')
    return bytes(out)


@app.route('/api/admin/user-analytics', methods=['GET'])
@admin_required
def admin_user_analytics():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT COALESCE(NULLIF(trim(city), ''), '(غير محدد)') AS c, COUNT(*) AS n
           FROM users GROUP BY c ORDER BY n DESC'''
    )
    by_city = [{'city': r[0], 'count': r[1]} for r in cursor.fetchall()]
    cursor.execute(
        '''SELECT COALESCE(NULLIF(trim(home_type), ''), '(غير محدد)') AS h, COUNT(*) AS n
           FROM users GROUP BY h ORDER BY n DESC'''
    )
    by_home = [{'home_type': r[0], 'count': r[1]} for r in cursor.fetchall()]
    cursor.execute('SELECT COUNT(DISTINCT user_id) FROM plants')
    users_with_plants = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM users')
    total_users = cursor.fetchone()[0]
    cursor.execute(
        '''SELECT u.id, u.name, u.city, COUNT(p.id) AS plant_count
           FROM users u LEFT JOIN plants p ON p.user_id = u.id
           GROUP BY u.id ORDER BY plant_count DESC LIMIT 30'''
    )
    top_by_plants = [{'user_id': r[0], 'name': r[1], 'city': r[2], 'plant_count': r[3]} for r in cursor.fetchall()]
    conn.close()
    return jsonify({
        'total_users': total_users,
        'users_with_at_least_one_plant': users_with_plants,
        'by_city': by_city,
        'by_home_type': by_home,
        'top_users_by_plant_count': top_by_plants,
        'data_sources': {
            'city_and_home': 'حقول `city` و `home_type` في جدول `users` — تُملأ عند التسجيل أو من لوحة المدير.',
            'plants_per_user': 'يُحسب من جدول `plants` عبر `user_id`.',
            'care_history': 'جدول `care_logs` مرتبط بكل نبتة عبر `plant_id` ثم بالمستخدم.',
            'no_gps': 'لا يوجد تتبع موقع تلقائي؛ المدينة قيمة نصية يختارها المستخدم أو الإدارة.',
            'database_file': os.path.basename(DATABASE),
        },
    })


@app.route('/api/admin/report-export', methods=['GET'])
@admin_required
def admin_report_export():
    fmt = (request.args.get('format') or 'json').lower().strip()
    bundle = _fetch_admin_report_bundle()
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    base = 'florabit-report-%s' % stamp

    if fmt in ('json', 'application/json'):
        data = json.dumps(bundle, ensure_ascii=False, indent=2)
        return Response(
            data,
            mimetype='application/json; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s.json"' % base},
        )

    if fmt in ('txt', 'text'):
        raw = _bundle_to_txt(bundle)
        return Response(
            raw,
            mimetype='text/plain; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s.txt"' % base},
        )

    if fmt == 'csv':
        raw = _bundle_to_csv_bytes(bundle)
        return Response(
            raw.encode('utf-8'),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s.csv"' % base},
        )

    if fmt in ('sql', 'sqlite'):
        raw = _bundle_to_sql(bundle, mysql_style=False)
        return Response(
            raw,
            mimetype='text/plain; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s.sql"' % base},
        )

    if fmt in ('mysql', 'mariadb'):
        raw = _bundle_to_sql(bundle, mysql_style=True)
        return Response(
            raw,
            mimetype='text/plain; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s-mysql.sql"' % base},
        )

    if fmt in ('xlsx', 'excel'):
        blob = _bundle_to_xlsx_bytes(bundle)
        if blob is None:
            return jsonify({'error': 'ثبّت الحزمة: pip install openpyxl'}), 501
        return Response(
            blob,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="%s.xlsx"' % base},
        )

    if fmt == 'pdf':
        blob = _bundle_to_pdf_bytes(bundle)
        if blob is None:
            return jsonify({'error': 'ثبّت الحزمة: pip install fpdf2'}), 501
        return Response(
            blob,
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename="%s.pdf"' % base},
        )

    return jsonify({'error': 'صيغة غير معروفة. جرّب: json, txt, csv, sql, mysql, xlsx, pdf'}), 400


@app.route('/api/plants/upcoming-care', methods=['GET'])
def upcoming_care():
    user_id = request.args.get('user_id', type=int)
    conn = get_db()
    cursor = conn.cursor()
    if user_id:
        cursor.execute('SELECT * FROM plants WHERE user_id = ?', (user_id,))
    else:
        if session.get('role') != 'admin':
            conn.close()
            return jsonify({'error': 'حدّد user_id أو سجّل دخول كمدير'}), 403
        cursor.execute('SELECT * FROM plants')
    plants = [dict(row) for row in cursor.fetchall()]
    conn.close()
    now = datetime.now()
    items = []
    for p in plants:
        wid = int(p.get('watering_interval_days') or 7)
        fid = int(p.get('fertilizing_interval_days') or 30)
        t_lw = _parse_dt(p.get('last_watering_date'))
        t_lf = _parse_dt(p.get('last_fertilizing_date'))
        next_w = (t_lw + timedelta(days=wid)) if t_lw else (now + timedelta(days=wid))
        next_f = (t_lf + timedelta(days=fid)) if t_lf else (now + timedelta(days=fid))
        items.append({
            'plant_id': p['id'],
            'name': p['name'],
            'next_watering': next_w.strftime('%Y-%m-%d %H:%M:%S'),
            'next_fertilizing': next_f.strftime('%Y-%m-%d %H:%M:%S'),
            'watering_overdue': next_w <= now,
            'fertilizing_overdue': next_f <= now,
            'watering_days_until': max(0, (next_w - now).days),
            'fertilizing_days_until': max(0, (next_f - now).days),
        })
    return jsonify(items)


@app.route('/api/user/smart-summary', methods=['GET'])
def smart_summary():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'user_id مطلوب'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM plants WHERE user_id = ?', (user_id,))
    plants = [dict(row) for row in cursor.fetchall()]
    now = datetime.now()
    plants_out = []
    scores = []
    overdue_w = 0
    overdue_f = 0
    for p in plants:
        wid = int(p.get('watering_interval_days') or 7)
        fid = int(p.get('fertilizing_interval_days') or 30)
        t_lw = _parse_dt(p.get('last_watering_date'))
        t_lf = _parse_dt(p.get('last_fertilizing_date'))
        next_w = (t_lw + timedelta(days=wid)) if t_lw else (now + timedelta(days=wid))
        next_f = (t_lf + timedelta(days=fid)) if t_lf else (now + timedelta(days=fid))
        if next_w <= now:
            overdue_w += 1
        if next_f <= now:
            overdue_f += 1
        hs = _plant_health_score(p, now)
        scores.append(hs)
        plants_out.append({
            'id': p['id'],
            'name': p['name'],
            'type': p.get('type'),
            'health_score': hs,
            'status_ar': _health_status_ar(hs),
        })
    overall = int(round(sum(scores) / len(scores))) if scores else 100
    streak = _care_streak_days(conn, user_id)
    conn.close()
    return jsonify({
        'overall_health_score': overall,
        'overdue_watering_plants': overdue_w,
        'overdue_fertilizing_plants': overdue_f,
        'care_streak_days': streak,
        'tip_of_day': get_tip_of_day(),
        'plants': plants_out,
    })


@app.route('/api/care-logs', methods=['GET'])
def get_care_logs():
    plant_id = request.args.get('plant_id', type=int)
    if not plant_id:
        if session.get('role') != 'admin':
            return jsonify({'error': 'حدّد plant_id أو سجّل دخول كمدير'}), 403
    conn = get_db()
    cursor = conn.cursor()
    if plant_id:
        cursor.execute('SELECT * FROM care_logs WHERE plant_id = ? ORDER BY action_date DESC', (plant_id,))
    else:
        cursor.execute('SELECT * FROM care_logs ORDER BY action_date DESC')
    logs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(logs)


@app.route('/api/care-logs', methods=['POST'])
def create_care_log():
    data = request.get_json()
    if not data or not data.get('plant_id') or not data.get('action_type'):
        return jsonify({'error': 'plant_id and action_type required'}), 400
    if data['action_type'] not in ('watering', 'fertilizing', 'lighting'):
        return jsonify({'error': 'action_type must be watering, fertilizing, or lighting'}), 400
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(
        'INSERT INTO care_logs (plant_id, action_type, action_date, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
        (data['plant_id'], data['action_type'], now, data.get('notes', ''), now, now),
    )
    conn.commit()
    log_id = cursor.lastrowid
    if data['action_type'] == 'watering':
        cursor.execute('UPDATE plants SET last_watering_date = ? WHERE id = ?', (now, data['plant_id']))
    elif data['action_type'] == 'fertilizing':
        cursor.execute('UPDATE plants SET last_fertilizing_date = ? WHERE id = ?', (now, data['plant_id']))
    conn.commit()
    conn.close()
    return jsonify({'id': log_id, 'action_type': data['action_type'], 'action_date': now}), 201


@app.route('/api/reminders', methods=['GET', 'POST'])
def reminders_api():
    if request.method == 'GET':
        plant_id = request.args.get('plant_id', type=int)
        user_id = request.args.get('user_id', type=int)
        conn = get_db()
        cursor = conn.cursor()
        try:
            cursor.execute('SELECT 1 FROM reminders LIMIT 1')
        except sqlite3.OperationalError:
            conn.close()
            return jsonify([])
        if plant_id:
            denied = _plant_access_denied(plant_id)
            if denied is not None:
                conn.close()
                return denied
            cursor.execute(
                'SELECT * FROM reminders WHERE plant_id = ? ORDER BY due_date',
                (plant_id,),
            )
        elif user_id:
            sid = session.get('user_id')
            role = session.get('role')
            hdr = request.headers.get('X-User-Id')
            ok = role == 'admin' or sid == user_id or (
                hdr is not None and str(hdr) == str(user_id)
            )
            if not ok:
                conn.close()
                return jsonify({'error': 'غير مصرّح'}), 403
            cursor.execute(
                '''SELECT r.* FROM reminders r
                   JOIN plants p ON r.plant_id = p.id
                   WHERE p.user_id = ? ORDER BY r.due_date''',
                (user_id,),
            )
        else:
            conn.close()
            return jsonify({'error': 'حدّد plant_id أو user_id'}), 400
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify(rows)
    data = request.get_json() or {}
    pid = data.get('plant_id')
    due = data.get('due_date')
    if not pid or not due:
        return jsonify({'error': 'plant_id و due_date مطلوبان'}), 400
    rtype = data.get('reminder_type') or 'other'
    if rtype not in ('watering', 'fertilizing', 'other', 'lighting'):
        rtype = 'other'
    denied = _plant_access_denied(int(pid))
    if denied is not None:
        return denied
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        cursor.execute(
            '''INSERT INTO reminders (plant_id, reminder_type, due_date, is_sent, notes, created_at, updated_at)
               VALUES (?, ?, ?, 0, ?, ?, ?)''',
            (pid, rtype, due, data.get('notes', '') or '', now, now),
        )
        conn.commit()
        rid = cursor.lastrowid
    except sqlite3.OperationalError as e:
        conn.close()
        return jsonify({'error': str(e)}), 500
    conn.close()
    return jsonify({'id': rid}), 201


@app.route('/api/reminders/<int:rid>', methods=['DELETE'])
def delete_reminder(rid):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT plant_id FROM reminders WHERE id = ?', (rid,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'غير موجود'}), 404
    denied = _plant_access_denied(row[0])
    if denied is not None:
        conn.close()
        return denied
    cursor.execute('DELETE FROM reminders WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# Flutter Web (نفس النطاق): ضع مخرجات `flutter build web --base-href /app/` داخل مجلد `flutter_web/`
FLUTTER_WEB_DIR = os.path.join(os.path.dirname(__file__), 'flutter_web')


@app.route('/app')
def flutter_web_redirect():
    return redirect('/app/', code=302)


@app.route('/app/', defaults={'path': ''})
@app.route('/app/<path:path>')
def flutter_web(path):
    root = FLUTTER_WEB_DIR
    index_path = os.path.join(root, 'index.html')
    if not os.path.isfile(index_path):
        return (
            '<!DOCTYPE html><html lang="ar"><meta charset="utf-8"><title>فلورابيت</title>'
            '<p>واجهة التطبيق (Flutter Web) غير مرفوعة بعد. من مجلد <code>app/app</code> نفّذ:</p>'
            '<pre>flutter build web --base-href /app/</pre>'
            '<p>ثم انسخ كل محتويات <code>build/web/</code> إلى <code>web/flutter_web/</code> وادفع للمستودع.</p>',
            503,
            {'Content-Type': 'text/html; charset=utf-8'},
        )
    if path:
        candidate = os.path.join(root, path)
        if os.path.isfile(candidate):
            return send_from_directory(root, path)
    return send_from_directory(root, 'index.html')


# Serve HTML pages
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/<path:path>')
def serve_page(path):
    if path.endswith('.html') or '.' not in path:
        return send_from_directory('.', path if path.endswith('.html') else f'{path}.html')
    return send_from_directory('.', path)


init_db()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
